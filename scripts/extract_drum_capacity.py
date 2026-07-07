# Extract per-drum cable capacity (meters) from packing list files.
# Sources ('Detail PL' sheet of each PGU-DE-*.xlsx):
#   - Meter rows: Quantity col = meters, unit 'M/метр' (Power/Control/I&C/FGSS/HRSG)
#   - EA rows with meters embedded in description ', 1800m' (AIS)
# Drum tag: Tag No. column if present, else package no. minus 'PGU-DE-XXX-' prefix
# (both match cable-material.json drumList format).
# Output: cable-dashboard/public/drum-capacity.json  { tag: { m, pk } }
import json, re
from pathlib import Path
from openpyxl import load_workbook

PACK_DIR = Path('cable-dashboard/public/files/packing')
OUT = Path('cable-dashboard/public/drum-capacity.json')

DRAWING_RE = re.compile(r'^(CCP-|WD[_-])', re.I)
DESC_METER_PATTERNS = [
  re.compile(r',\s*([\d,]+)\s*m\s*$', re.I),      # '..., 1800m'
  re.compile(r'-(\d+)\s*m\s*$', re.I),            # '...-13m'
  re.compile(r'Mtr\.?\s*([\d,]+)', re.I),         # '...Mtr.2000'
]
PKG_PREFIX_RE = re.compile(r'^PGU-DE-\d+-', re.I)

result = {}
stats = []

for f in sorted(PACK_DIR.glob('*.xlsx')):
    pk = f.stem
    wb = load_workbook(f, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if 'detail' in s.lower()), None)
    if not sheet:
        stats.append((pk, 0)); continue
    count = 0
    for row in wb[sheet].iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue
        pkg, desc, qty, unit = row[1], row[2], row[4], row[5]
        if not (isinstance(pkg, str) and pkg.upper().startswith('PGU-DE-')):
            continue
        qty = float(qty) if isinstance(qty, (int, float)) else (float(qty) if isinstance(qty, str) and qty.replace('.', '', 1).isdigit() else None)
        unit = (unit or '') if isinstance(unit, str) else ''
        meters = None
        meter_unit = qty and unit.strip().upper().startswith('M')
        if meter_unit:
            meters = qty
        elif isinstance(desc, str):
            for pat in DESC_METER_PATTERNS:
                m = pat.search(desc.strip())
                if m:
                    meters = float(m.group(1).replace(',', '')) * (qty or 1)
                    break
        if not meters:
            continue
        # tag: meter-unit rows may have a Tag No. column (c12..c18, skip drawing nos);
        # EA rows (AIS/DCS style) always derive from the package number.
        tag = None
        if meter_unit:
            for j in range(12, min(len(row), 19)):
                v = row[j]
                if (isinstance(v, str) and v.strip() and not DRAWING_RE.match(v.strip())
                        and ' ' not in v.strip() and not re.fullmatch(r'[\d-]+', v.strip())):
                    tag = v.strip()
                    break
        if not tag:
            tag = PKG_PREFIX_RE.sub('', pkg.strip())
        if tag in result:
            result[tag]['m'] += meters
            if pk not in result[tag]['pk']:
                result[tag]['pk'] += f' + {pk}'
        else:
            result[tag] = {'m': meters, 'pk': pk}
        count += 1
    stats.append((pk, count))

OUT.write_text(json.dumps(result, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
print(f'{len(result)} drums, {sum(v["m"] for v in result.values()):,.0f} m -> {OUT}')
for pk, n in stats:
    print(f'  {pk:20s} {n} drum rows')

# validate against cable-material drumList
mat = json.load(open('cable-dashboard/public/cable-material.json', encoding='utf-8'))
all_tags = {d for r in mat for d in (r.get('drumList') or [])}
hit = sum(1 for t in all_tags if t in result)
print(f'coverage: {hit}/{len(all_tags)} material drum tags have capacity')
missing = sorted(t for t in all_tags if t not in result)
print('missing sample:', missing[:15])
